"""
generate_review.py
------------------
Reads a chapter's raw input Excel, processes vocab and passages,
and produces a review Excel for the teacher to check pinyin and context sentences.

Rev 5 (M2.5 T2):
    The review workbook now preserves THREE sheets:
      - Meta       (copied verbatim from raw)
      - Content    (copied verbatim from raw)
      - Review     (generated: Traditional + English + Pinyin + ContextSentence)
    The Instructions sheet from the raw template is dropped in the review file.

    This is achieved by loading the raw workbook, dropping Instructions,
    adding the Review sheet, and saving under the _review.xlsx filename.
    All original formatting (Meta!B2 text format, Content!A wrap_text,
    comments, styling) is preserved automatically.
"""

import argparse
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from pypinyin import pinyin, Style

# Unicode-escaped Chinese punctuation to avoid chat-stripping issues
CHINESE_PERIOD = "\u3002"     # 。
CHINESE_EXCLAIM = "\uff01"    # ！
CHINESE_QUESTION = "\uff1f"   # ？
SENTENCE_DELIMS = CHINESE_PERIOD + CHINESE_EXCLAIM + CHINESE_QUESTION
BU_CHAR = "\u4e0d"            # 不


# ------------------------------------------------------------------
# Text / parsing helpers (unchanged from previous version)
# ------------------------------------------------------------------
def clean_text(text):
    """Strip whitespace AND non-breaking spaces from a string."""
    if text is None:
        return ""
    return str(text).replace("\xa0", "").strip()


def concat_column(sheet, column_letter, start_row=2):
    """Read all non-empty cells in a column starting from start_row."""
    values = []
    for row_cells in sheet.iter_rows(min_row=start_row):
        for cell in row_cells:
            if cell.column_letter == column_letter and cell.value is not None:
                cleaned = clean_text(cell.value)
                if cleaned:
                    values.append(cleaned)
    return values


def parse_meta(wb):
    """Extract ChapterID and ChapterTitle. Handles int (Excel auto-convert) or str.

    With T1's text-formatted Meta!B2, ChapterID should always arrive as a string
    like "00", "01", ... — but we keep the int/float safety net in case someone
    resaves the raw file and accidentally loses the '@' format.
    """
    meta = wb["Meta"]
    chapter_id = None
    chapter_title = None
    for row in meta.iter_rows(min_row=2, values_only=True):
        field = row[0]
        value = row[1] if len(row) > 1 else None
        if field == "ChapterID":
            if isinstance(value, (int, float)):
                chapter_id = f"{int(value):02d}"
            elif value is not None:
                chapter_id = clean_text(value).zfill(2)
        elif field == "ChapterTitle":
            chapter_title = clean_text(value)
    return chapter_id, chapter_title


def generate_pinyin_for_word(word):
    """Generate pinyin: Style.TONE, space-separated, base tones only (strip 不 sandhi)."""
    syllables = pinyin(word, style=Style.TONE)
    flat = [s[0] for s in syllables]
    # Strip 不 sandhi: force bù (base) instead of bú (pypinyin's auto-sandhi)
    if BU_CHAR in word:
        for i, char in enumerate(word):
            if char == BU_CHAR and i < len(flat) and flat[i] == "b\u00fa":
                flat[i] = "b\u00f9"
    return " ".join(flat)


def extract_context_sentence(word, passage):
    """Find first sentence in passage containing the word (split by 。！？)."""
    if not word or not passage:
        return ""
    pattern = f"[^{SENTENCE_DELIMS}]+[{SENTENCE_DELIMS}]"
    sentences = re.findall(pattern, passage)
    # Any trailing text without punctuation
    remaining = re.sub(pattern, "", passage).strip()
    if remaining:
        sentences.append(remaining)
    for sentence in sentences:
        if word in sentence:
            return sentence.strip()
    return ""


def sort_vocab_by_appearance(vocab_list, english_list, passage):
    """Sort by first-appearance in passage. Not-found words appended at end."""
    with_pos = []
    not_found = []
    for word, meaning in zip(vocab_list, english_list):
        pos = passage.find(word)
        if pos == -1:
            not_found.append((word, meaning))
        else:
            with_pos.append((pos, word, meaning))
    with_pos.sort(key=lambda x: x[0])
    sorted_vocab = [w for _, w, _ in with_pos] + [w for w, _ in not_found]
    sorted_english = [e for _, _, e in with_pos] + [e for _, e in not_found]
    not_found_words = [w for w, _ in not_found]
    return sorted_vocab, sorted_english, not_found_words


# ------------------------------------------------------------------
# Core chapter processing (returns the LOADED workbook too)
# ------------------------------------------------------------------
def process_chapter(input_path):
    print(f"[INFO] Loading: {input_path}")
    if not os.path.exists(input_path):
        print(f"[ERROR] File not found: {input_path}")
        sys.exit(1)

    wb = load_workbook(input_path)

    chapter_id, chapter_title = parse_meta(wb)
    if not chapter_id:
        print("[ERROR] ChapterID is missing in Meta sheet.")
        sys.exit(1)
    print(f"[INFO] ChapterID: {chapter_id}")
    print(f"[INFO] ChapterTitle: {chapter_title}")

    content = wb["Content"]
    passage_rows = concat_column(content, "A")
    vocab_rows = concat_column(content, "B")
    english_rows = concat_column(content, "C")

    # NOTE: For pinyin + context extraction we still need a single string
    # to search within. We join all passage rows with '\n' as before.
    # The ORIGINAL row structure (one passage per row, with Alt+Enter
    # paragraph breaks preserved) stays intact in the Content sheet
    # itself — which is what generate_chapter.py will read in T3.
    full_passage_for_search = "\n".join(passage_rows)

    vocab_joined = "|".join(vocab_rows)
    english_joined = "|".join(english_rows)
    vocab_list = [clean_text(v) for v in vocab_joined.split("|") if clean_text(v)]
    english_list = [clean_text(e) for e in english_joined.split("|") if clean_text(e)]

    print(f"[INFO] Passage rows: {len(passage_rows)}")
    print(f"[INFO] Passage length (concatenated for search): "
          f"{len(full_passage_for_search)} characters")
    print(f"[INFO] Vocab items: {len(vocab_list)}")
    print(f"[INFO] English items: {len(english_list)}")

    if len(vocab_list) != len(english_list):
        print(f"[ERROR] Vocab count ({len(vocab_list)}) doesn't match "
              f"English count ({len(english_list)}).")
        sys.exit(1)

    vocab_list, english_list, not_found = sort_vocab_by_appearance(
        vocab_list, english_list, full_passage_for_search
    )
    if not_found:
        print(f"[WARN] {len(not_found)} vocab word(s) not found in passage:")
        for w in not_found:
            print(f"       - {w}  (add context sentence manually)")

    print(f"[INFO] Generating pinyin and extracting context sentences...")
    results = []
    for word, meaning in zip(vocab_list, english_list):
        results.append({
            "traditional": word,
            "english": meaning,
            "pinyin": generate_pinyin_for_word(word),
            "context": extract_context_sentence(word, full_passage_for_search),
        })

    return wb, chapter_id, chapter_title, results


# ------------------------------------------------------------------
# NEW: add a Review sheet to the already-loaded raw workbook,
# and drop Instructions. Meta + Content sheets are preserved verbatim.
# ------------------------------------------------------------------
def add_review_sheet(wb, chapter_id, chapter_title, results):
    """Modify wb in place:
       - Drop Instructions sheet (if present).
       - Add a Review sheet with Traditional / English / Pinyin / Context.
    """
    # Drop Instructions sheet — not needed in the review file
    if "Instructions" in wb.sheetnames:
        del wb["Instructions"]

    # If a Review sheet already exists (e.g., re-running), replace it
    if "Review" in wb.sheetnames:
        del wb["Review"]

    ws = wb.create_sheet("Review")

    # Title + subtitle
    ws["A1"] = f"Chapter {chapter_id}: {chapter_title}"
    ws["A1"].font = Font(bold=True, size=14, color="2B6CB0")
    ws.merge_cells("A1:D1")

    ws["A2"] = ("REVIEW: check Pinyin (fix multi-pronunciation errors) "
                "and ContextSentence (fix if wrong or missing).")
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:D2")

    # Header row
    headers = [
        "Traditional",
        "English",
        "Pinyin (auto - review!)",
        "ContextSentence (auto - review!)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4A5568")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for i, r in enumerate(results, start=4):
        ws.cell(row=i, column=1, value=r["traditional"])
        ws.cell(row=i, column=2, value=r["english"])
        ws.cell(row=i, column=3, value=r["pinyin"])
        ws.cell(row=i, column=4, value=r["context"])
        if not r["context"]:
            ws.cell(row=i, column=4).fill = PatternFill("solid", fgColor="FFF3CD")
            ws.cell(row=i, column=4).comment = Comment(
                "Word not found in passage. Please add context sentence manually.",
                "System",
            )

    # Column widths + freeze panes
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 70
    ws.freeze_panes = "A4"


# ------------------------------------------------------------------
# Entry point
# ------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Generate review Excel from raw chapter input."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Path to chapterXX_raw.xlsx (relative to content-pipeline/)",
    )
    args = parser.parse_args()

    input_path = args.input
    if not os.path.isabs(input_path):
        input_path = os.path.join("content-pipeline", input_path)

    wb, chapter_id, chapter_title, results = process_chapter(input_path)

    # Modify the loaded workbook: drop Instructions, add Review sheet
    add_review_sheet(wb, chapter_id, chapter_title, results)

    # Write out under the _review.xlsx name
    input_dir = os.path.dirname(input_path)
    input_basename = os.path.basename(input_path)
    output_basename = input_basename.replace("_raw.xlsx", "_review.xlsx")
    output_path = os.path.join(input_dir, output_basename)
    wb.save(output_path)

    print()
    print("=" * 60)
    print(f"[OK] Review file created: {output_path}")
    print(f"     Sheets in review file: {wb.sheetnames}")
    print(f"     Words processed: {len(results)}")
    missing_ctx = sum(1 for r in results if not r["context"])
    if missing_ctx > 0:
        print(f"     Words needing manual context sentence: {missing_ctx}")
    print("=" * 60)
    print()
    print("Next steps:")
    print("  1. Download the review file to your PC.")
    print("  2. Verify pinyin (fix multi-pronunciation errors) and context sentences.")
    print("  3. Save as chapterXX_final.xlsx and re-upload to Codespace.")


if __name__ == "__main__":
    main()
