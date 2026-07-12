"""
generate_chapter.py
-------------------
Orchestrator: reads _final.xlsx, produces chapter JSON + audio MP3s,
updates the master chapters index.

Rev 5 (M2.5 T3):
    - Reads Meta directly from the Meta sheet (always present after M2.5 T2).
    - Reads passages directly from the Content sheet, Column A.
      One passage per row. Internal \n (Alt+Enter) preserved verbatim.
    - Reads vocab data from the Review sheet.
    - Removes the old fallback that reconstructed passages from context
      sentences (no longer needed).
    - Fails fast with a clear error if Meta / Content / Review are missing.

Usage:
    python content-pipeline/generate_chapter.py --input inputs/chapter00_final.xlsx
    python content-pipeline/generate_chapter.py --input inputs/chapter00_final.xlsx --passage-audio-only
"""

import argparse
import asyncio
import json
import os
import sys

from openpyxl import load_workbook

# We still delegate audio generation to generate_audio.py.
# Meta / passage / vocab reading uses shared helpers from generate_common.py.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_audio import generate_chapter_audio, generate_passage_audio
from generate_common import clean_text, parse_meta, validate_book_and_paths, LANGUAGES, to_simplified


def read_passages(wb):
    """Read passages from the Content sheet, Column A.
    One passage per row. Internal \n (from Alt+Enter) preserved."""
    if "Content" not in wb.sheetnames:
        print("[ERROR] Workbook is missing the Content sheet.")
        sys.exit(1)

    content = wb["Content"]
    passages = []
    for row in content.iter_rows(min_row=2, values_only=True):
        cell = row[0] if row else None
        if cell is None:
            continue
        cleaned = clean_text(cell)
        if cleaned:
            passages.append(cleaned)
    return passages


def read_review_vocab(wb):
    """Read vocab rows from the Review sheet.
    Data rows start at row 4 (title r1, subtitle r2, header r3)."""
    if "Review" not in wb.sheetnames:
        print("[ERROR] Workbook is missing the Review sheet.")
        sys.exit(1)

    ws = wb["Review"]
    words = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row and row[0]:
            words.append({
                "traditional": clean_text(row[0]),
                "english": clean_text(row[1]) if len(row) > 1 and row[1] else "",
                "pinyin": clean_text(row[2]) if len(row) > 2 and row[2] else "",
                "context": clean_text(row[3]) if len(row) > 3 and row[3] else "",
            })
    return words


# ------------------------------------------------------------------
# Build JSON
# ------------------------------------------------------------------
def build_chapter_json(book_id, chapter_id, chapter_title, passages, words):
    """Build the chapter JSON in the schema our HTML app expects (rev 7, M8.3c-1.5).

    Schema (M8.3c-1.5 adds simplified variants of chapter title, passages,
    vocab words, and context sentences):
    {
      "bookId": "01",
      "chapterId": "01",
      "chapterTitle": "...",
      "chapterTitleSimplified": "...",
      "passages": [
        "Passage 1 paragraph 1\nPassage 1 paragraph 2",
        "Passage 2 text"
      ],
      "passagesSimplified": [
        "Simplified version of passage 1",
        "Simplified version of passage 2"
      ],
      "vocab": [
        {
          "id": "b01_ch01_w001",
          "traditional": "...",
          "simplified": "...",
          "pinyin": "...",
          "english": "...",
          "contextSentence": "...",
          "contextSentenceSimplified": "...",
          "audio": {
            "mandarin": {
              "isolated": "audio/book01/chapter01/b01_ch01_w001_m_isolated.mp3",
              "sentence": "audio/book01/chapter01/b01_ch01_w001_m_sentence.mp3"
            },
            "cantonese": {
              "isolated": "audio/book01/chapter01/b01_ch01_w001_c_isolated.mp3",
              "sentence": "audio/book01/chapter01/b01_ch01_w001_c_sentence.mp3"
            }
          }
        }
      ]
    }
    """
    vocab = []
    audio_folder = f"audio/book{book_id}/chapter{chapter_id}"
    for i, w in enumerate(words, start=1):
        word_id = f"b{book_id}_ch{chapter_id}_w{i:03d}"

        # Build nested audio structure with all configured languages
        audio_dict = {}
        for lang_name, lang_config in LANGUAGES.items():
            suffix = lang_config["suffix"]
            audio_dict[lang_name] = {
                "isolated": f"{audio_folder}/{word_id}_{suffix}_isolated.mp3",
            }
            if w["context"]:
                audio_dict[lang_name]["sentence"] = f"{audio_folder}/{word_id}_{suffix}_sentence.mp3"

        entry = {
            "id": word_id,
            "traditional": w["traditional"],
            "simplified": to_simplified(w["traditional"]),
            "english": w.get("english", ""),
            "pinyin": w.get("pinyin", ""),
            "contextSentence": w["context"],
            "contextSentenceSimplified": to_simplified(w["context"]),
            "audio": audio_dict,
        }
        vocab.append(entry)

    # Simplified versions of chapter title and passages (M8.3c-1.5 extension)
    passages_simplified = [to_simplified(p) for p in passages]

    return {
        "bookId": book_id,
        "chapterId": chapter_id,
        "chapterTitle": chapter_title,
        "chapterTitleSimplified": to_simplified(chapter_title),
        "passages": passages,
        "passagesSimplified": passages_simplified,
        "vocab": vocab,
    }


def update_chapters_index(book_id, chapter_id, chapter_title, word_count):
    """Add or update this chapter in data/book{book_id}/chapters-index.json.

    Each book has its own per-book chapters index (M8.3c-1).
    """
    book_dir = f"data/book{book_id}"
    index_path = f"{book_dir}/chapters-index.json"
    os.makedirs(book_dir, exist_ok=True)

    if os.path.exists(index_path):
        with open(index_path, "r", encoding="utf-8") as f:
            index = json.load(f)
    else:
        index = {"chapters": []}

    existing = None
    for ch in index["chapters"]:
        if ch["id"] == chapter_id:
            existing = ch
            break

    entry = {"id": chapter_id, "title": chapter_title, "wordCount": word_count}

    if existing:
        idx = index["chapters"].index(existing)
        index["chapters"][idx] = entry
        action = "Updated"
    else:
        index["chapters"].append(entry)
        index["chapters"].sort(key=lambda x: x["id"])
        action = "Added"

    with open(index_path, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

    return action


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------
async def main_async(input_path, passage_audio_only):
    print(f"[INFO] Processing: {input_path}")
    print()

    if not os.path.exists(input_path):
        print(f"[ERROR] File not found: {input_path}")
        sys.exit(1)

    wb = load_workbook(input_path)

    book_id, chapter_id, chapter_title = parse_meta(wb)
    validate_book_and_paths(book_id, chapter_id, input_path)
    passages = read_passages(wb)
    words = read_review_vocab(wb)

    print(f"[INFO] BookID:      {book_id}")
    print(f"[INFO] ChapterID:   {chapter_id}")
    print(f"[INFO] Title:       {chapter_title}")
    print(f"[INFO] Passages:    {len(passages)}")
    print(f"[INFO] Vocab words: {len(words)}")
    print()

    # ---- M9.2 fast path: --passage-audio-only ----
    # Regenerate passage MP3s and merge the passageAudio block into an
    # existing chapterYY.json. Vocab audio and vocab JSON are NOT touched.
    if passage_audio_only:
        book_dir = f"data/book{book_id}"
        json_path = f"{book_dir}/chapter{chapter_id}.json"

        if not os.path.exists(json_path):
            print(f"[ERROR] {json_path} not found.")
            print(f"        Run without --passage-audio-only first to build the full chapter JSON.")
            sys.exit(1)

        print("[INFO] --passage-audio-only mode: regenerating passage audio only.")
        print("[INFO] Vocab audio and vocab JSON will NOT be touched.")
        print()

        passages_simp = [to_simplified(p) for p in passages]
        passage_audio = await generate_passage_audio(
            book_id, chapter_id, passages, passages_simp
        )

        with open(json_path, "r", encoding="utf-8") as f:
            chapter_json = json.load(f)
        chapter_json["passageAudio"] = passage_audio
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(chapter_json, f, ensure_ascii=False, indent=2)

        print()
        print("=" * 60)
        print(f"[OK] Merged passageAudio block into {json_path}")
        print("=" * 60)
        return


    # ---- Build JSON ----
    print("[STEP 1] Building chapter JSON...")
    chapter_json = build_chapter_json(book_id, chapter_id, chapter_title, passages, words)
    book_dir = f"data/book{book_id}"
    json_path = f"{book_dir}/chapter{chapter_id}.json"
    os.makedirs(book_dir, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(chapter_json, f, ensure_ascii=False, indent=2)
    print(f"         Written: {json_path}")

    # ---- Update chapters index ----
    print()
    print("[STEP 2] Updating chapters index...")
    action = update_chapters_index(book_id, chapter_id, chapter_title, len(words))
    print(f"         {action} chapter {chapter_id} in {book_dir}/chapters-index.json")

    # ---- Generate vocab audio ----
    print()
    print(f"[STEP 3] Generating bilingual vocab audio ({len(LANGUAGES)} languages)...")
    print(f"         (This takes ~60s for a small chapter with both Mandarin + Cantonese)")
    await generate_chapter_audio(input_path)

    # ---- M9.2: Passage audio + JSON merge ----
    print()
    print(f"[STEP 4] Generating passage audio ({len(passages)} passages x {len(LANGUAGES)} languages)...")
    passages_simp = [to_simplified(p) for p in passages]
    passage_audio = await generate_passage_audio(
        book_id, chapter_id, passages, passages_simp
    )

    # Re-open the JSON written in STEP 1 and merge the passageAudio block.
    with open(json_path, "r", encoding="utf-8") as f:
        chapter_json = json.load(f)
    chapter_json["passageAudio"] = passage_audio
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(chapter_json, f, ensure_ascii=False, indent=2)
    print(f"         Merged passageAudio block into {json_path}")

    # ---- Summary ----
    print()
    print("=" * 60)
    print(f"[OK] Book {book_id} Chapter {chapter_id} published!")
    print()
    print("Files produced:")
    print(f"  - {json_path}")
    print(f"  - {book_dir}/chapters-index.json (updated)")
    print(f"  - audio/book{book_id}/chapter{chapter_id}/ (4 MP3s per word + 2 MP3s per passage)")
    print()
    print("Next steps:")
    print(f"  1. Verify JSON: cat {json_path}")
    print(f"  2. Commit + push in Codespace")
    print(f"  3. Wait ~1 min for GitHub Pages to rebuild")
    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(
        description="Publish a chapter: generate JSON + audio, update index."
    )
    parser.add_argument("--input", required=True, help="Path to chapterXX_final.xlsx")

    parser.add_argument(
        "--passage-audio-only",
        action="store_true",
        help="Only regenerate passage audio (M9) and merge into existing chapterYY.json. "
            "Vocab audio and vocab JSON left untouched. Used for backfilling old chapters."
    )

    args = parser.parse_args()

    input_path = args.input
    if not os.path.isabs(input_path):
        input_path = os.path.join("content-pipeline", input_path)

    asyncio.run(main_async(input_path, args.passage_audio_only))


if __name__ == "__main__":
    main()