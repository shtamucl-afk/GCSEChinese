"""
generate_audio.py
-----------------
Generate MP3 audio files from a chapter's _final.xlsx using edge-tts.

Produces up to 2 MP3s per vocab word:
  - {word}_isolated.mp3   - just the word
  - {word}_sentence.mp3   - the context sentence

Voice: zh-CN-XiaoxiaoNeural
Rate:  -10% (slightly slower for learner comprehension)

Runs only in GitHub Codespaces (see Project Plan Section 11.8).

Rev 5 (M2.5 T4):
    - Reads ChapterID + ChapterTitle directly from the Meta sheet.
    - Filename-regex inference is retained only as a last-resort fallback
      (in case a resave accidentally strips the Meta sheet).
    - Reads vocab rows from the Review sheet as before.

Usage:
    python content-pipeline/generate_audio.py --input inputs/chapter00_final.xlsx
"""

import argparse
import asyncio
import os
import tempfile
import sys

from openpyxl import load_workbook
import edge_tts

# Shared pipeline helpers (M8.3c-1)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_common import (
    clean_text,
    parse_meta,
    validate_book_and_paths,
    LANGUAGES,
    split_passage_into_sentences,
    mp3_duration_ms,
)


def read_final_excel(input_path):
    """Read _final.xlsx. Returns (book_id, chapter_id, chapter_title, words list)."""
    wb = load_workbook(input_path)
    book_id, chapter_id, chapter_title = parse_meta(wb)
    validate_book_and_paths(book_id, chapter_id, input_path)

    # Vocab data still lives on the Review sheet
    if "Review" not in wb.sheetnames:
        print("[ERROR] Workbook is missing the Review sheet.")
        sys.exit(1)
    ws = wb["Review"]

    words = []
    # Data rows start at row 4 (title r1, subtitle r2, header r3)
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row and row[0]:  # has traditional character
            traditional = clean_text(row[0])
            context = clean_text(row[3]) if len(row) > 3 and row[3] else ""
            words.append({"traditional": traditional, "context": context})

    return book_id, chapter_id, chapter_title, words


# ------------------------------------------------------------------
# Audio generation
# ------------------------------------------------------------------
async def generate_mp3(text, output_path, voice, rate):
    """Generate one MP3 file using edge-tts."""
    tts = edge_tts.Communicate(text, voice, rate=rate)
    await tts.save(output_path)


async def generate_chapter_audio(input_path):
    print(f"[INFO] Loading: {input_path}")
    if not os.path.exists(input_path):
        print(f"[ERROR] File not found: {input_path}")
        sys.exit(1)

    book_id, chapter_id, chapter_title, words = read_final_excel(input_path)
    print(f"[INFO] BookID:    {book_id}")
    print(f"[INFO] ChapterID: {chapter_id}")
    print(f"[INFO] Title:     {chapter_title}")
    print(f"[INFO] Words:     {len(words)}")
    print(f"[INFO] Languages: {list(LANGUAGES.keys())}")

    audio_dir = f"audio/book{book_id}/chapter{chapter_id}"
    os.makedirs(audio_dir, exist_ok=True)
    print(f"[INFO] Output folder: {audio_dir}/")
    print()

    # Count total MP3s to generate:
    # for each word: 1 isolated MP3 per language, plus 1 sentence MP3 per language if context exists
    n_langs = len(LANGUAGES)
    total = 0
    for w in words:
        total += n_langs  # isolated in each language
        if w["context"]:
            total += n_langs  # sentence in each language

    count = 0

    for i, w in enumerate(words, start=1):
        word = w["traditional"]
        context = w["context"]
        word_id = f"b{book_id}_ch{chapter_id}_w{i:03d}"

        for lang_name, lang_config in LANGUAGES.items():
            voice = lang_config["voice"]
            rate = lang_config["rate"]
            suffix = lang_config["suffix"]

            # Isolated word MP3
            count += 1
            iso_filename = f"{word_id}_{suffix}_isolated.mp3"
            iso_path = os.path.join(audio_dir, iso_filename)
            print(f"[{count}/{total}] {iso_filename} ({word}, {lang_name}) ... ",
                  end="", flush=True)
            try:
                await generate_mp3(word, iso_path, voice, rate)
                size = os.path.getsize(iso_path)
                print(f"OK ({size} bytes)")
            except Exception as e:
                print(f"FAILED: {e}")

            # Sentence MP3 (only if context exists)
            if context:
                count += 1
                sent_filename = f"{word_id}_{suffix}_sentence.mp3"
                sent_path = os.path.join(audio_dir, sent_filename)
                print(f"[{count}/{total}] {sent_filename} ({word}, {lang_name}) ... ",
                      end="", flush=True)
                try:
                    await generate_mp3(context, sent_path, voice, rate)
                    size = os.path.getsize(sent_path)
                    print(f"OK ({size} bytes)")
                except Exception as e:
                    print(f"FAILED: {e}")

        if not context:
            print(f"       (no sentence MP3 for {word} - no context)")

    print()
    print("=" * 60)
    print(f"[OK] Audio generation complete!")
    print(f"     Folder: {audio_dir}/")
    print(f"     Files created: {len(os.listdir(audio_dir))}")
    print(f"     Languages: {list(LANGUAGES.keys())}")
    print("=" * 60)

# ------------------------------------------------------------------
# Passage audio generation (M9.1)
# ------------------------------------------------------------------
async def generate_passage_audio(book_id, chapter_id, passages, passages_simplified=None):
    """Generate one MP3 per passage per language, plus per-sentence offsets.

    AApproach: synthesize per sentence, measure durations, concatenate:
      1. Split each passage into sentences on 。！？…； + '\\n' paragraph breaks.
      2. Synthesize each sentence individually to a temp MP3, measuring its
         real duration via mutagen.
      3. Concatenate the temp MP3s byte-wise into the final passage MP3.
      4. Compute sentence startMs/endMs from the cumulative measured durations.
      5. Clean up temp files.

    Note: sentence text passed to edge-tts is Traditional. Sentence *offsets*
    differ per language because Mandarin (Xiaoxiao) and Cantonese (WanLung)
    voices synthesize at different speeds. textSimplified is populated by
    splitting the Simplified passages the same way; OpenCC t2s is 1-to-1,
    so counts and positions match exactly.

    Returns: list of passageAudio entries matching the M9 JSON schema.
    """
    audio_dir = f"audio/book{book_id}/chapter{chapter_id}"
    os.makedirs(audio_dir, exist_ok=True)

    split_by_passage = [split_passage_into_sentences(p) for p in passages]
    split_simp = None
    if passages_simplified is not None:
        split_simp = [split_passage_into_sentences(p) for p in passages_simplified]
        # Sanity check — this should never fail, but fail loudly if it does.
        for i, (t, s) in enumerate(zip(split_by_passage, split_simp)):
            if len(t) != len(s):
                print(f"[ERROR] Passage {i+1}: Traditional split has {len(t)} "
                      f"sentences but Simplified split has {len(s)}. "
                      f"Check for stray terminator characters.")
                sys.exit(1)

    n_langs = len(LANGUAGES)
    total_units = sum(len(s) for s in split_by_passage) * n_langs
    count = 0

    print(f"[INFO] Generating passage audio: {len(passages)} passages "
          f"x {n_langs} languages")
    print(f"[INFO] Total sentence syntheses: {total_units}")
    print()

    result = []
    for p_idx, sentences in enumerate(split_by_passage):
        entry = {"passageIndex": p_idx}
        p_num = f"{p_idx + 1:02d}"

        for lang_name, lang_config in LANGUAGES.items():
            voice = lang_config["voice"]
            rate = lang_config["rate"]
            suffix = lang_config["suffix"]

            passage_filename = f"b{book_id}_ch{chapter_id}_p{p_num}_{suffix}.mp3"
            passage_path = os.path.join(audio_dir, passage_filename)

            with tempfile.TemporaryDirectory(
                prefix=f"passage_p{p_num}_{suffix}_"
            ) as tmp_dir:
                per_sentence = []  # list of (tmp_path, duration_ms, sentence_dict)
                for s_idx, sentence in enumerate(sentences):
                    count += 1
                    tmp_file = os.path.join(tmp_dir, f"s{s_idx:03d}.mp3")
                    print(f"[{count}/{total_units}] p{p_num} s{s_idx+1:02d}/"
                          f"{len(sentences):02d} ({lang_name}) ... ",
                          end="", flush=True)
                    try:
                        await generate_mp3(sentence["text"], tmp_file, voice, rate)
                        dur = mp3_duration_ms(tmp_file)
                        per_sentence.append((tmp_file, dur, sentence))
                        print(f"OK ({dur} ms)")
                    except Exception as e:
                        print(f"FAILED: {e}")
                        raise

                # Byte-wise concat. edge-tts MP3s are CBR from the same encoder,
                # so plain concatenation plays back cleanly in all browsers.
                with open(passage_path, "wb") as out_f:
                    for tmp_file, _, _ in per_sentence:
                        with open(tmp_file, "rb") as in_f:
                            out_f.write(in_f.read())

            # Authoritative duration of the concatenated file.
            total_ms = mp3_duration_ms(passage_path)

            sentences_out = []
            cursor = 0
            for s_idx, (_, dur, sentence) in enumerate(per_sentence):
                text_simp = ""
                if split_simp is not None:
                    text_simp = split_simp[p_idx][s_idx]["text"]
                sentences_out.append({
                    "index": s_idx,
                    "paragraph": sentence["paragraph"],
                    "startMs": cursor,
                    "endMs": cursor + dur,
                    "text": sentence["text"],
                    "textSimplified": text_simp,
                })
                cursor += dur

            # Align the final endMs to the concatenated total to absorb the
            # sub-millisecond rounding drift from summing individual durations.
            if sentences_out:
                sentences_out[-1]["endMs"] = total_ms

            entry[lang_name] = {
                "file": f"{audio_dir}/{passage_filename}",
                "durationMs": total_ms,
                "sentences": sentences_out,
            }

            size = os.path.getsize(passage_path)
            print(f"    -> {passage_filename} ({total_ms} ms, "
                  f"{size} bytes, {len(sentences_out)} sentences)")
            print()

        result.append(entry)

    print("=" * 60)
    print(f"[OK] Passage audio generation complete!")
    print(f"    Passages: {len(passages)}")
    print(f"    Files created: {len(passages) * n_langs}")
    print("=" * 60)

    return result


def main():
    parser = argparse.ArgumentParser(description="Generate MP3 audio via edge-tts.")
    parser.add_argument("--input", required=True, help="Path to chapterXX_final.xlsx")
    args = parser.parse_args()

    input_path = args.input
    if not os.path.isabs(input_path):
        input_path = os.path.join("content-pipeline", input_path)

    asyncio.run(generate_chapter_audio(input_path))


if __name__ == "__main__":
    main()