"""
build_raw_template.py
---------------------
Creates the blank Raw Input template that Fiona downloads to author each chapter.

Run ONCE (or whenever you want to regenerate a fresh blank template).
Not part of the regular chapter authoring workflow.

Output: content-pipeline/templates/Raw_Input_template.xlsx

Rev 5 notes (Milestone 2.5 - T1):
- Passages are now ONE PER ROW in Column A of the Content sheet.
- Alt+Enter within a cell creates paragraph breaks (\n) inside that passage.
- Each row becomes a separate entry in the JSON "passages" array.
- ChapterID cell (Meta!B2) is forced to TEXT format so "00" stays as "00".
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
import os


def create_raw_template():
    """
    Create the blank raw input template with the agreed structure:
      - Meta sheet: ChapterID + ChapterTitle
        (ChapterID is TEXT-formatted so "00", "01", ... are preserved)
      - Content sheet: 3 columns (Passage | Vocabulary | English)
        Passage = one passage per row; Alt+Enter for paragraph breaks.
        Vocabulary + English are independent lists using '|' as delimiter.
      - Instructions sheet: how-to guide for the teacher.
    """
    wb = Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Meta
    # ------------------------------------------------------------------
    meta = wb.active
    meta.title = "Meta"

    meta["A1"] = "Field"
    meta["B1"] = "Value"
    meta["A2"] = "ChapterID"
    meta["A3"] = "ChapterTitle"

    # Style header row
    for cell in [meta["A1"], meta["B1"]]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4A5568")
        cell.alignment = Alignment(horizontal="center")

    # --- Force Column B of Meta to TEXT format ---
    # '@' is Excel's built-in text format code.
    # This ensures "00", "01", "07" are stored/displayed as text,
    # not silently converted to numeric 0, 1, 7.
    for row_idx in range(2, 20):  # cover a generous range
        meta.cell(row=row_idx, column=2).number_format = "@"

    # Pre-seed B2 with an empty string so the format sticks even before typing.
    meta["B2"] = ""
    meta["B3"] = ""

    # Add hints as cell comments
    meta["A2"].comment = Comment(
        "Two-digit chapter number as TEXT, e.g., 00, 01, 02, ..., 20. "
        "Cell is formatted as text so leading zeros are preserved. "
        "Determines JSON filename (chapterXX.json) and audio folder (audio/chapterXX/).",
        "System",
    )
    meta["A3"].comment = Comment(
        "Chapter title in Chinese or English. Shown in the app dashboard.",
        "System",
    )

    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 60

    # ------------------------------------------------------------------
    # Sheet 2: Content
    # ------------------------------------------------------------------
    content = wb.create_sheet("Content")

    headers = ["Passage", "Vocabulary", "English"]
    for col_idx, header in enumerate(headers, start=1):
        cell = content.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4A5568")
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    content.column_dimensions["A"].width = 80
    content.column_dimensions["B"].width = 40
    content.column_dimensions["C"].width = 60

    # Row heights for readability
    content.row_dimensions[1].height = 25

    # Wrap text in Column A so Alt+Enter paragraph breaks are visible.
    for row_idx in range(2, 50):
        content.cell(row=row_idx, column=1).alignment = Alignment(
            wrap_text=True, vertical="top"
        )

    # Hints via comments
    content["A1"].comment = Comment(
        "PASSAGE column.\n"
        "One passage per row.\n"
        "If a passage contains multiple paragraphs, use Alt+Enter INSIDE the "
        "same cell to create paragraph breaks. Each Alt+Enter becomes a \\n "
        "in the JSON. Each row becomes a separate entry in the passages array.",
        "System",
    )
    content["B1"].comment = Comment(
        "VOCABULARY column. Traditional Chinese words separated by | (pipe). "
        "You can spread across multiple rows - Python concatenates. "
        "Example: 創建|相輔相成|思維",
        "System",
    )
    content["C1"].comment = Comment(
        "ENGLISH column. English meanings separated by | (pipe), in the SAME "
        "ORDER as Vocabulary. Total count in B and C must match after "
        "concatenation. Example: Establish|Complement each other|Mindset",
        "System",
    )

    # ------------------------------------------------------------------
    # Sheet 3: Instructions (for the human filling this in)
    # ------------------------------------------------------------------
    instr = wb.create_sheet("Instructions")

    instructions = [
        ("DFSNMGCSEChinese - Raw Input Template", True),
        ("", False),
        ("HOW TO FILL IN THIS TEMPLATE", True),
        ("", False),
        ("1. Save this file as: chapterXX_raw.xlsx (replace XX with chapter number)", False),
        ("   Example: chapter01_raw.xlsx", False),
        ("", False),
        ("2. Meta sheet:", False),
        ("   - ChapterID: two-digit text, e.g. 00, 01, 02, ...", False),
        ("     (Cell B2 is text-formatted so leading zeros are preserved.)", False),
        ("   - ChapterTitle: title shown in the app dashboard.", False),
        ("", False),
        ("3. Content sheet - three independent columns:", False),
        ("   Column A (Passage):", False),
        ("     - ONE PASSAGE PER ROW.", False),
        ("     - If a passage has multiple paragraphs, use Alt+Enter INSIDE the", False),
        ("       same cell to create paragraph breaks.", False),
        ("     - Each row becomes a separate entry in the JSON passages array.", False),
        ("     - Alt+Enter breaks become \\n in the JSON.", False),
        ("   Column B (Vocabulary): Chinese words separated by | (pipe). Multiple rows OK.", False),
        ("   Column C (English): English meanings separated by | (pipe), same order as B.", False),
        ("", False),
        ("4. IMPORTANT: The total count of items in B and C must match.", False),
        ("   Python will error clearly if counts don't align.", False),
        ("", False),
        ("5. Upload this file to your Codespace at:", False),
        ("   content-pipeline/inputs/chapterXX_raw.xlsx", False),
        ("", False),
        ("6. Run: python content-pipeline/generate_review.py --input inputs/chapterXX_raw.xlsx", False),
        ("", False),
        ("7. Download the generated chapterXX_review.xlsx, check the auto-generated pinyin", False),
        ("   and context sentences, correct any errors, save as chapterXX_final.xlsx", False),
        ("", False),
        ("8. Upload chapterXX_final.xlsx back to Codespace, then run publish script:", False),
        ("   python content-pipeline/generate_chapter.py --input inputs/chapterXX_final.xlsx", False),
        ("", False),
        ("PASSAGE STRUCTURE EXAMPLES", True),
        ("Example A - single passage, single paragraph:", False),
        ("   Row 2, Column A: 英華國際學校創建於1990年，設有小學和中學。", False),
        ("", False),
        ("Example B - single passage, multiple paragraphs (use Alt+Enter):", False),
        ("   Row 2, Column A: 英華國際學校創建於1990年。[Alt+Enter]雙語是這個學校的教學特色。", False),
        ("", False),
        ("Example C - multiple passages:", False),
        ("   Row 2, Column A: First passage text.", False),
        ("   Row 3, Column A: Second passage text.", False),
        ("   Row 4, Column A: Third passage text.", False),
        ("", False),
        ("DELIMITER RULE (Vocabulary + English columns)", True),
        ("Use | (pipe character) to separate items in Vocabulary and English columns.", False),
        ("Do NOT use commas - some English meanings naturally contain commas.", False),
        ("Example: park, garden|to like, prefer  <- would break if using comma delimiter", False),
        ("", False),
        ("MULTI-PRONUNCIATION CHARACTERS (多音字)", True),
        ("Some Traditional Chinese characters have multiple pronunciations that pypinyin", False),
        ("may guess incorrectly. Common examples:", False),
        ("  長: cháng (long) vs zhǎng (elder/grow) - e.g., 長方形 should be cháng fāng xíng", False),
        ("  行: xíng (walk) vs háng (row/profession) - e.g., 銀行 = yín háng", False),
        ("  樂: lè (happy) vs yuè (music) - e.g., 音樂 = yīn yuè", False),
        ("Fix these manually during the review step.", False),
    ]

    for row_idx, (text, is_bold) in enumerate(instructions, start=1):
        cell = instr.cell(row=row_idx, column=1, value=text)
        if is_bold:
            cell.font = Font(bold=True, size=12, color="2B6CB0")

    instr.column_dimensions["A"].width = 100

    return wb


def main():
    os.makedirs("content-pipeline/templates", exist_ok=True)
    wb = create_raw_template()
    output_path = "content-pipeline/templates/Raw_Input_template.xlsx"
    wb.save(output_path)

    print(f"[OK] Blank raw template created: {output_path}")
    print()
    print("Rev 5 changes applied:")
    print("  - Meta!B2 (ChapterID) is now TEXT-formatted (leading zeros preserved).")
    print("  - Content!A: one passage per row; Alt+Enter for paragraph breaks.")
    print("  - Instructions sheet updated with passage-structure examples.")
    print()
    print("Next steps:")
    print("  1. Download this file from Codespace to your PC.")
    print("  2. Fill it in with your chapter content.")
    print("  3. Save as chapterXX_raw.xlsx and upload to content-pipeline/inputs/")
    print("  4. Run generate_review.py to produce the review file.")


if __name__ == "__main__":
    main()
